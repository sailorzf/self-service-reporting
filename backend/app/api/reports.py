import uuid
import io
import urllib.parse
from datetime import datetime, timedelta, date
from decimal import Decimal
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse, HTMLResponse
from sqlalchemy.orm import Session
from openpyxl import Workbook
from sqlalchemy import text
from app.database import get_db
from app.models import Report, DataType
from app.schemas import ReportCreate, ReportResponse, QueryExecute, QueryResult, ReportConfig
from app.report_engine import ReportEngine, DataFormatter
from app.export_html import generate_html_report

router = APIRouter()


def serialize_excel_cell(v):
    if isinstance(v, (datetime, date)):
        return v
    if isinstance(v, Decimal):
        return float(v)
    return v


@router.post("/", response_model=ReportResponse)
def create_report(report: ReportCreate, db: Session = Depends(get_db)):
    # Canvas reports may have no data_type_id
    if report.data_type_id:
        dt = db.query(DataType).filter(DataType.id == report.data_type_id).first()
        if not dt:
            raise HTTPException(404, "数据表不存在")
    config = report.config_json
    if isinstance(config, dict):
        pass  # already a dict
    elif config is None:
        config = {}
    else:
        config = config.model_dump()
    r = Report(name=report.name, data_type_id=report.data_type_id, config_json=config)
    db.add(r)
    db.commit()
    db.refresh(r)
    return r


@router.put("/{report_id}", response_model=ReportResponse)
def update_report(report_id: int, report: ReportCreate, db: Session = Depends(get_db)):
    r = db.query(Report).filter(Report.id == report_id).first()
    if not r:
        raise HTTPException(404, "报表不存在")
    r.name = report.name
    r.data_type_id = report.data_type_id
    config = report.config_json
    if isinstance(config, dict):
        r.config_json = config
    elif config is not None:
        r.config_json = config.model_dump()
    db.commit()
    db.refresh(r)
    return r

@router.get("/", response_model=list[ReportResponse])
def list_reports(db: Session = Depends(get_db)):
    return db.query(Report).order_by(Report.updated_at.desc()).all()

@router.post("/execute", response_model=QueryResult)
def execute_query(query: QueryExecute, db: Session = Depends(get_db)):
    from sqlalchemy import text
    from app.config import settings

    # Raw SQL path
    if query.raw_sql:
        sql = query.raw_sql.strip().rstrip(';')
        if not sql.upper().startswith("SELECT"):
            raise HTTPException(403, "只允许执行SELECT语句")
        if "LIMIT" not in sql.upper():
            sql = f"{sql} LIMIT {getattr(settings, 'max_result_rows', 1000)}"
        try:
            result = db.execute(text(sql))
            rows = result.fetchall()
            headers = list(result.keys())
        except Exception as e:
            raise HTTPException(400, f"SQL执行失败: {str(e)}")
        chart_data = DataFormatter.to_chart(headers, rows) if rows else None
        return QueryResult(
            headers=headers,
            rows=[list(r) for r in rows],
            chart_data=chart_data
        )

    # Existing logic
    engine = ReportEngine(db)
    if query.config and query.config.tables and len(query.config.tables) > 1:
        table_map = {}
        for t in query.config.tables:
            dt = db.query(DataType).filter(DataType.id == t.data_type_id).first()
            if not dt:
                raise HTTPException(404, f"数据表 {t.data_type_id} 不存在")
            db_prefix = f"{dt.database_name}." if dt.database_name else ""
            table_map[t.alias] = f"{db_prefix}{dt.table_name}"
        return engine.execute_multi(query.config, table_map)
    if query.data_type_id and query.config:
        dt = db.query(DataType).filter(DataType.id == query.data_type_id).first()
        if not dt:
            raise HTTPException(404, "数据表不存在")
        return engine.execute(query.config, dt.table_name, dt.database_name)
    raise HTTPException(400, "需要指定data_type_id、tables或raw_sql")

@router.post("/{report_id}/execute", response_model=QueryResult)
def execute_report(report_id: int, db: Session = Depends(get_db)):
    r = db.query(Report).filter(Report.id == report_id).first()
    if not r:
        raise HTTPException(404, "报表不存在")
    dt = db.query(DataType).filter(DataType.id == r.data_type_id).first()
    if not dt:
        raise HTTPException(404, "数据表不存在")
    engine = ReportEngine(db)
    config = ReportConfig(**r.config_json)
    return engine.execute(config, dt.table_name, dt.database_name)

@router.post("/{report_id}/share")
def share_report(report_id: int, days: int = Query(default=7), db: Session = Depends(get_db)):
    r = db.query(Report).filter(Report.id == report_id).first()
    if not r:
        raise HTTPException(404, "报表不存在")
    r.shared_token = str(uuid.uuid4())
    r.token_expires = datetime.now() + timedelta(days=days)
    db.commit()
    return {"share_url": f"/share/{r.shared_token}", "expires": r.token_expires}

@router.get("/{report_id}/export")
def export_report(report_id: int, db: Session = Depends(get_db)):
    r = db.query(Report).filter(Report.id == report_id).first()
    if not r:
        raise HTTPException(404, "报表不存在")

    config = r.config_json

    # Canvas format
    if isinstance(config, dict) and "canvas" in config and "components" in config:
        wb = Workbook()
        for comp in config["components"]:
            if comp.get("type") not in ("table", "bar", "line", "pie"):
                continue
            sql = comp.get("sql", "").strip().rstrip(";")
            if not sql:
                continue
            try:
                result = db.execute(text(sql))
                rows = result.fetchall()
                headers = list(result.keys())
                ws = wb.create_sheet(title=comp.get("name", comp.get("type", "data"))[:31])
                ws.append(headers)
                for row in rows:
                    ws.append([serialize_excel_cell(v) for v in row])
            except Exception:
                pass
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        if not wb.sheetnames:
            ws = wb.create_sheet(title="数据")
            ws.append(["该报表没有可导出的组件"])
    else:
        # Legacy format
        dt = db.query(DataType).filter(DataType.id == r.data_type_id).first()
        if not dt:
            raise HTTPException(404, "数据表不存在")
        engine = ReportEngine(db)
        result = engine.execute(ReportConfig(**config), dt.table_name, dt.database_name)
        wb = Workbook()
        ws = wb.active
        ws.append(result["headers"])
        for row in result["rows"]:
            ws.append([serialize_excel_cell(v) for v in row])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename_encoded = urllib.parse.quote(r.name)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename_encoded}"}
    )


@router.get("/{report_id}/export/html")
def export_report_html(report_id: int, db: Session = Depends(get_db)):
    r = db.query(Report).filter(Report.id == report_id).first()
    if not r:
        raise HTTPException(404, "报表不存在")
    html_content = generate_html_report(r, db)
    return HTMLResponse(content=html_content)
